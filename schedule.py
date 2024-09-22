import csv
import pandas as pd 
import openpyxl
from openpyxl import load_workbook 
import users
import time
from users import Teacher, Student 
import match
import main_calendar
import assignment

INF = 1e9
class Session: 
    """
    """
    def __init__(self, ID:str, students, teachers):
        self.ID = '' 
        self.dates = ''
        self.students = [student for student in students if student.Session == ID]
        self.teachers = [teacher for teacher in teachers if ID in teacher.AvailabilityData]
        self.optimal_schedule = {}
    
class Schedule: 
    """
    Builds a schedule by solving an instance of the unbalanced assignment problem
    """
    def __init__(self, students, teachers): 
        self.students = students 
        self.teachers = teachers
        self.enum_students = enumerate(students)
        self.enum_teachers = enumerate(teachers)
        self.dist_matrix = self.build_matrix(self.students, self.teachers)
        self.matches = []

    def build_matrix(self, students, teachers): 
        """
        """
        schedule_by_OSRM = [[student.OSRM_get_distance_to(teacher[1]) for teacher in enum_teachers] for student in enum_students]
        schedule_by_OX = [[student.OSRM_get_distance_to(teacher[1]) for teacher in enum_teachers] for student in enum_students]
        return matrix
        
    
    def SCIPY_get_matches(self):
        solver = ScipyLinearSumAssignment(self.dist_matrix)
        matches = solver.get_matches()
        return matches

def get_teacher_data(survey_responses:str):
    """
    takes as input survey responses in csv file
    assumes survey responses are organized as follows: 
        ID | FirstName | LastName | EmailAddress | Address | Availability
    assumes Availability is a string of comma-separated date ranges, like this: "April 1 - 30 2024, May 1 - 30 2024"
    returns a dictionary mapping a teacher to sessions they can teach 
    """
    teachers = {}
    wb = load_workbook(filename = survey_responses, data_only=True) 
    ws = wb['teacher']
    for row in ws.iter_rows(ws.min_row+1, 4):
        time.sleep(1)
        ID = row[0].value
        First = row[1].value
        Last = row[2].value
        ClinicName = row[3].value
        ClinicAddress = row[4].value
        AvailableDates = row[5].value
        AvailabilityData = [main_calendar[date] for date in AvailableDates.split(',')]
        teacher = Teacher(ID, First, Last, ClinicName, ClinicAddress, AvailabilityData)
        teachers[ID] = teacher 
    return teachers

def get_student_data(survey_responses:str):
    """
    takes as input survey responses in csv file
    assumes survey responses are organized as follows
        ID | FirstName | LastName | Address |TravelMethods | Session#
    """
    students = {}
    wb = load_workbook(filename = survey_responses, data_only=True) 
    ws = wb['student']
    for row in ws.iter_rows(ws.min_row+1, 4):
        time.sleep(1)
        ID = row[0].value
        First = row[1].value
        Last = row[2].value
        Address = row[3].value
        TravelMethod = row[4].value
        Session = row[5].value
        student = Student(ID, First, Last, Address, TravelMethod, Session)
        students[ID] = student
    return students
    
def make_student_schedules(survey_responses:dict): 
    """
    takes as input geocoded survey responses in dict 
    returns a student schedule  
    """
    pass

def main():
    file1 = "test_teacher_data.xlsx"
    file2 = "test_student_data.xlsx"

    teacher_records = get_teacher_data(file1)
    student_records = get_student_data(file2)
    for teacher in teacher_records:
        teacher_records[teacher].print_teacher()

if __name__=="__main__": 
    main()